from openpyxl import load_workbook

from services.file.app.parser.base_parser import BaseParser


class XlsxParser(BaseParser):
    async def extract_text(self, file_path: str) -> str:
        workbook = load_workbook(
            filename=file_path,
            data_only=True
        )

        lines = []

        for worksheets in workbook.worksheets:
            lines.append(f"Sheets: {worksheets.title}")

            for rows in worksheets.iter_rows(values_only=True):
                values = [
                    str(cell) for cell in rows if cell is not None
                ]
                if values:
                    lines.append(" | ".join(values))

            lines.append("")

        workbook.close()

        return "\n".join(lines)
